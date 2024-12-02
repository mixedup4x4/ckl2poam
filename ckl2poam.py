#!/usr/bin/env python3
# ckl2poam.py
# Converts STIG Viewer checklist file to a POA&M Excel spreadsheet
# Last updated by Allyn Stott (allyn.stott@navy.mil)

import sys
from xml.dom.minidom import parseString
from openpyxl import Workbook

VERSION = "1.0"


def show_help():
    """Display usage instructions."""
    print(f"ckl2poam version {VERSION}")
    print(f"Usage: {sys.argv[0]} input.ckl output.xlsx")
    sys.exit(1)


def get_attribute_data(node, index, default=""):
    """Safely retrieve data from ATTRIBUTE_DATA nodes."""
    try:
        return node.getElementsByTagName("ATTRIBUTE_DATA")[index].firstChild.data
    except (IndexError, AttributeError):
        return default


def get_text_content(node, tag, default=""):
    """Safely retrieve text content of a specific tag."""
    try:
        return node.getElementsByTagName(tag)[0].firstChild.data
    except (IndexError, AttributeError):
        return default


def main():
    if len(sys.argv) < 3:
        show_help()

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    # Read and parse the input file
    try:
        with open(input_file, 'r') as file:
            data = file.read()
            dom = parseString(data)
    except Exception as e:
        print(f"Error reading input file: {e}")
        sys.exit(1)

    # Initialize Excel workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "POA&M"
    ws2 = wb.create_sheet(title="Remediation")

    # Define column headers
    poam_headers = [
        'Weakness', 'CAT', 'Severity', 'IA Control and Impact Code', 'POC',
        'Resources Required', 'Scheduled Completion Date', 'Milestones with Completion Dates',
        'Milestone Changes', 'Source Identifying Weakness', 'Status', 'Finding Details', 'Comments'
    ]
    remediation_headers = [
        'Weakness', 'CAT', 'Severity', 'IA Control and Impact Code', 'Source Identifying Weakness',
        'Vulnerability Discussion', 'Check Content', 'Fix Text', 'Finding Details', 'Comments'
    ]

    # Add headers to worksheets
    for col, header in enumerate(poam_headers, start=1):
        ws1.cell(row=1, column=col, value=header)

    for col, header in enumerate(remediation_headers, start=1):
        ws2.cell(row=1, column=col, value=header)

    # Process vulnerabilities
    row_counter = 2
    vulnerabilities = dom.getElementsByTagName('VULN')

    for vuln in vulnerabilities:
        status = get_text_content(vuln, "STATUS")

        if status in ("Not_Reviewed", "Open"):
            severity = get_attribute_data(vuln, 1)
            cat = {"high": "CAT I", "medium": "CAT II", "low": "CAT III"}.get(severity, "")

            # Extract data fields
            rule_title = get_attribute_data(vuln, 5)
            group_title = get_attribute_data(vuln, 2)
            rule_id = get_attribute_data(vuln, 3)
            vuln_discussion = get_attribute_data(vuln, 6)
            ia_controls = get_attribute_data(vuln, 7)
            chk_content = get_attribute_data(vuln, 8)
            fix_text = get_attribute_data(vuln, 9)
            finding_details = get_text_content(vuln, "FINDING_DETAILS")
            comments = get_text_content(vuln, "COMMENTS")

            # Construct rows for each worksheet
            poam_values = [
                rule_title, cat, severity, ia_controls, '', '', '', '', '',
                f"{group_title}\n{rule_id}", status, finding_details, comments
            ]
            remediation_values = [
                rule_title, cat, severity, ia_controls, f"{group_title}\n{rule_id}",
                vuln_discussion, chk_content, fix_text, finding_details, comments
            ]

            # Write rows to Excel sheets
            for col, value in enumerate(poam_values, start=1):
                ws1.cell(row=row_counter, column=col, value=value)

            for col, value in enumerate(remediation_values, start=1):
                ws2.cell(row=row_counter, column=col, value=value)

            row_counter += 1

    # Save the workbook
    try:
        wb.save(output_file)
        print(f"Excel file created successfully: {output_file}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        sys.exit(1)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit("Aborted by user request.")

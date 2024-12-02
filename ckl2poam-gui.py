#!/usr/bin/env python3
# ckl2poam.py
# Converts STIG Viewer checklist file to a POA&M Excel spreadsheet
# Modernized and updated for Python 3

import sys
from tkinter import Tk, Frame, Button, Entry, filedialog
from xml.dom.minidom import parseString
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

VERSION = "1.0"


class CKL2POAMGUI:

    def __init__(self):
        self.fin = None
        self.fout = None
        self.dom = None

        app = Tk()
        app.title('ckl2poam GUI')

        # Frame 1: Input file selection
        fr1 = Frame(app, width=300, height=100)
        fr1.pack(side="top", pady=10)
        get_file_button = Button(fr1, text='Open CKL file ...', command=self.get_file)
        get_file_button.pack(side='left', padx=5)
        self.filein = Entry(fr1, width=40)
        self.filein.pack(side='right', padx=5)

        # Frame 2: Output file selection
        fr2 = Frame(app, width=300, height=100)
        fr2.pack(side="top", pady=10)
        save_file_button = Button(fr2, text='Save XLSX file ...', command=self.save_file)
        save_file_button.pack(side='left', padx=5)
        self.fileout = Entry(fr2, width=40)
        self.fileout.pack(side='right', padx=5)

        # Frame 3: Action buttons
        fr3 = Frame(app, width=300, height=100)
        fr3.pack(side="bottom", pady=10)
        okay_button = Button(fr3, text='Generate', command=self.create_xlsx)
        okay_button.pack(side="left", padx=10)
        cancel_button = Button(fr3, text='Cancel', command=self.kill_app)
        cancel_button.pack(side="right", padx=10)

        # Center the GUI window
        app.geometry('450x200+400+200')
        app.mainloop()

    def kill_app(self):
        sys.exit(0)

    def get_file(self):
        self.fin = filedialog.askopenfilename(filetypes=[("Checklist Files", "*.ckl"), ("All Files", "*.*")])
        if self.fin:
            self.filein.delete(0, 'end')
            self.filein.insert(0, self.fin)

    def save_file(self):
        self.fout = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if self.fout:
            self.fileout.delete(0, 'end')
            self.fileout.insert(0, self.fout)

    def create_xlsx(self):
        if not self.fin or not self.fout:
            print("Error: Both input and output files must be specified!")
            return

        try:
            with open(self.fin, 'r') as file:
                data = file.read()
                self.dom = parseString(data)
        except Exception as e:
            print(f"Error reading input file: {e}")
            return

        try:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = 'POA&M'
            ws2 = wb.create_sheet(title='Remediation')

            # Column headers
            poam_headers = ['Weakness', 'CAT', 'Severity', 'IA Control and Impact Code', 'POC',
                            'Resources Required', 'Scheduled Completion Date', 'Milestones with Completion Dates',
                            'Milestone Changes', 'Source Identifying Weakness', 'Status', 'Finding Details', 'Comments']
            remediation_headers = ['Weakness', 'CAT', 'Severity', 'IA Control and Impact Code', 'Source Identifying Weakness',
                                   'Vulnerability Discussion', 'Check Content', 'Fix Text', 'Finding Details', 'Comments']

            # Write headers
            for col_num, header in enumerate(poam_headers, start=1):
                ws1.cell(row=1, column=col_num, value=header)

            for col_num, header in enumerate(remediation_headers, start=1):
                ws2.cell(row=1, column=col_num, value=header)

            # Populate data
            row_counter = 2
            for vuln in self.dom.getElementsByTagName('VULN'):
                status = vuln.getElementsByTagName('STATUS')[0].firstChild.data
                if status in ('Not_Reviewed', 'Open'):
                    # Extract values with error handling
                    def get_value(tag, index, default=''):
                        try:
                            return vuln.getElementsByTagName(tag)[index].firstChild.data
                        except (IndexError, AttributeError):
                            return default

                    severity = get_value('ATTRIBUTE_DATA', 1)
                    cat = {'high': 'CAT I', 'medium': 'CAT II', 'low': 'CAT III'}.get(severity, '')

                    rule_title = get_value('ATTRIBUTE_DATA', 5)
                    group_title = get_value('ATTRIBUTE_DATA', 2)
                    rule_id = get_value('ATTRIBUTE_DATA', 3)
                    vuln_discussion = get_value('ATTRIBUTE_DATA', 6)
                    ia_controls = get_value('ATTRIBUTE_DATA', 7)
                    chk_content = get_value('ATTRIBUTE_DATA', 8)
                    fix_text = get_value('ATTRIBUTE_DATA', 9)
                    finding_details = get_value('FINDING_DETAILS', 0)
                    comments = get_value('COMMENTS', 0)

                    poam_values = [rule_title, cat, severity, ia_controls, '', '', '', '', '',
                                   f"{group_title}\n{rule_id}", status, finding_details, comments]
                    remediation_values = [rule_title, cat, severity, ia_controls,
                                          f"{group_title}\n{rule_id}", vuln_discussion,
                                          chk_content, fix_text, finding_details, comments]

                    # Write rows
                    for col_num, value in enumerate(poam_values, start=1):
                        ws1.cell(row=row_counter, column=col_num, value=value)

                    for col_num, value in enumerate(remediation_values, start=1):
                        ws2.cell(row=row_counter, column=col_num, value=value)

                    row_counter += 1

            wb.save(self.fout)
            print(f"Excel file created successfully: {self.fout}")

        except Exception as e:
            print(f"Error creating Excel file: {e}")


if __name__ == '__main__':
    try:
        CKL2POAMGUI()
    except KeyboardInterrupt:
        sys.exit("Aborted by user request.")
